import requests
import pandas as pd 
import aiohttp
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
from aiogram import types
from aiogram import Dispatcher
from aiogram.fsm.context import FSMContext
from aiogram.filters import Command, StateFilter
from aiogram.types import FSInputFile
from aiogram.fsm.state import State, StatesGroup
from keyboard import menu_keyboard  
from config import API_URL, REPORT_ALL_URL, REPORT_API_URL


class ExpenseForm(StatesGroup):
    waiting_for_title = State()  
    waiting_for_date = State()  
    waiting_for_amount = State()  

class ReportForm(StatesGroup):
    waiting_for_start_date = State()  
    waiting_for_end_date = State()    

class DelExpenseForm(StatesGroup):
    waiting_for_expense_id = State()  
    waiting_for_confirmation = State()  


class EditExpenseForm(StatesGroup):
    waiting_for_expense_id = State()  
    waiting_for_new_title = State()  
    waiting_for_new_amount = State()  

async def start_cmd(message: types.Message):
    await message.answer("Привіт! Я бот для обліку витрат. Обери дію:", reply_markup=menu_keyboard)


async def add_expense(message: types.Message, state: FSMContext):
    await message.answer("Введіть назву витрат (наприклад: 'Щомісячна оплата за інтернет'):")
    await state.set_state(ExpenseForm.waiting_for_title)  

async def get_expense_title(message: types.Message, state: FSMContext):
    title = message.text
    await state.update_data(title=title)  
    await message.answer("Введіть дату витрат у форматі dd.mm.YYYY:")
    await state.set_state(ExpenseForm.waiting_for_date)

async def get_expense_date(message: types.Message, state: FSMContext):
    date = message.text
    await state.update_data(date=date)  
    await message.answer("Введіть суму витрат (в UAH):")
    await state.set_state(ExpenseForm.waiting_for_amount)

async def get_expense_amount(message: types.Message, state: FSMContext):
    try:
        amount_uah = float(message.text)  
    except ValueError:
        await message.answer("Помилка: Сума має бути числом.")
        return

    data = await state.get_data()  
    title = data.get("title")
    date = data.get("date")

    expense_data = {
        "title": title,
        "date": date,
        "amount_uah": amount_uah
    }

    response = requests.post(API_URL, json=expense_data)

    if response.status_code == 200:
        await message.answer("Витрата успішно додана!")
    else:
        await message.answer("Помилка при додаванні витрат.")

    await message.answer("Обери наступну дію:", reply_markup=menu_keyboard)
    await state.clear()  



async def get_report(message: types.Message, state: FSMContext):
    await message.answer("Введіть початкову дату звіту (формат: dd.mm.YYYY):")
    await state.set_state(ReportForm.waiting_for_start_date)


async def get_start_date(message: types.Message, state: FSMContext):
    start_date = message.text
    try:
        
        datetime.strptime(start_date, "%d.%m.%Y")
    except ValueError:
        await message.answer("Помилка: Невірний формат дати. Будь ласка, використовуйте формат dd.mm.YYYY.")
        return

    
    await state.update_data(start_date=start_date)
    
    
    await message.answer("Введіть кінцеву дату звіту (формат: dd.mm.YYYY):")
    await state.set_state(ReportForm.waiting_for_end_date)


async def get_end_date(message: types.Message, state: FSMContext):
    end_date = message.text
    try:
        
        datetime.strptime(end_date, "%d.%m.%Y")
    except ValueError:
        await message.answer("Помилка: Невірний формат дати. Будь ласка, використовуйте формат dd.mm.YYYY.")
        return

    
    await state.update_data(end_date=end_date)

    
    data = await state.get_data()
    start_date = data["start_date"]
    end_date = data["end_date"]

    
    response = requests.get(f"{REPORT_API_URL}?start_date={start_date}&end_date={end_date}")
    
    if response.status_code == 200:
        
        with open("report.xlsx", "wb") as file:
            file.write(response.content)

        await message.answer("Звіт успішно отримано!")

        
        await message.answer_document(FSInputFile("report.xlsx"), caption="Ваш отчет о расходах за указанный период.")
    else:
        await message.answer("Помилка при отриманні звіту.")

    
    await state.clear()



async def edit_expense(message: types.Message, state: FSMContext):
    await message.answer("Ми готові редагувати всі витрати. Будь ласка, зачекайте, поки ми отримаємо звіт з усіма витратами.")
    
    
    response = requests.get(f"{REPORT_ALL_URL}")
    
    
    if response.status_code == 200:
        with open("all_expenses.xlsx", "wb") as file:
            file.write(response.content)

        
        await message.answer("Усі витрати успішно отримані! Обери статтю для редагування, надіславши її ID:")

        
        await message.answer_document(FSInputFile("all_expenses.xlsx"), caption="Список усіх витрат з їх ID.")
        
        
        await state.set_state(EditExpenseForm.waiting_for_expense_id)
    else:
        await message.answer("Помилка при отриманні всіх витрат.")

async def get_expense_id(message: types.Message, state: FSMContext):
    expense_id = message.text
    await state.update_data(expense_id=expense_id)
    
  
    async with aiohttp.ClientSession() as session:
        async with session.get(f"{REPORT_ALL_URL}") as resp:  
            if resp.status == 200:
                content_type = resp.headers.get('Content-Type', '')
                
                if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                    
                    excel_data = await resp.read()  
                    wb = load_workbook(BytesIO(excel_data))
                    sheet = wb.active
                    
                    
                    expenses = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):  
                        expenses.append({
                            'id': row[0],
                            'title': row[1],
                            'date': row[2],
                            'amount_uah': row[3],
                            'amount_usd': row[4]
                        })

                    
                    expense_data = next((expense for expense in expenses if expense['id'] == int(expense_id)), None)

                    if expense_data:
                        title = expense_data['title']
                        amount_uah = expense_data['amount_uah']
                        await message.answer(
                            f"Дані про статтю:\nНазва: {title}\nСума: {amount_uah} UAH\nВведіть нову назву статті:"
                        )
                        
                        await state.set_state(EditExpenseForm.waiting_for_new_title)
                    else:
                        await message.answer("Стаття з таким ID не знайдена. Спробуйте знову.")
                else:
                    await message.answer("Отримано несподіваний формат даних. Спробуйте ще раз.")
            else:
                await message.answer("Помилка при отриманні даних. Спробуйте ще раз.")





async def get_new_title(message: types.Message, state: FSMContext):
    new_title = message.text
    await state.update_data(new_title=new_title)
    await message.answer("Введіть нову суму статті (в UAH):")
    await state.set_state(EditExpenseForm.waiting_for_new_amount)


async def get_new_amount(message: types.Message, state: FSMContext):
    try:
        new_amount_uah = float(message.text)
    except ValueError:
        await message.answer("Помилка: Сума має бути числом.")
        return

    data = await state.get_data()
    expense_id = data["expense_id"]
    new_title = data["new_title"]

   
    expense_data = {
        "title": new_title,
        "amount_uah": new_amount_uah
    }

    response = requests.put(f"{API_URL}/{expense_id}", json=expense_data)

    if response.status_code == 200:
        await message.answer("Стаття успішно оновлена!")
    else:
        await message.answer("Помилка при оновленні статті.")

    await message.answer("Обери наступну дію:", reply_markup=menu_keyboard)
    await state.clear()


async def del_expense(message: types.Message, state: FSMContext):
    response = requests.get(f"{REPORT_ALL_URL}")
    
    
    if response.status_code == 200:
        with open("all_expenses.xlsx", "wb") as file:
            file.write(response.content)

        
        await message.answer_document(FSInputFile("all_expenses.xlsx"), caption="Список всех расходов с их ID.")
        
        
        await state.set_state(EditExpenseForm.waiting_for_expense_id)
    else:
        await message.answer("Помилка при отриманні всіх витрат.")
    await message.answer("Введіть ID витрат, які хочете видалити:")
    await state.set_state(DelExpenseForm.waiting_for_expense_id)


async def delete_expense(message: types.Message, state: FSMContext):
    expense_id = message.text
    await state.update_data(expense_id=expense_id)
    
    
    async with aiohttp.ClientSession() as session:
        async with session.get(f"{REPORT_ALL_URL}") as resp:  
            if resp.status == 200:
                content_type = resp.headers.get('Content-Type', '')
                
                if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                    
                    excel_data = await resp.read()  
                    wb = load_workbook(BytesIO(excel_data))
                    sheet = wb.active
                    
                    
                    expenses = []
                    for row in sheet.iter_rows(min_row=2, values_only=True): 
                        expenses.append({
                            'id': row[0],
                            'title': row[1],
                            'date': row[2],
                            'amount_uah': row[3],
                            'amount_usd': row[4]
                        })

                    
                    expense_data = next((expense for expense in expenses if expense['id'] == int(expense_id)), None)

                    if expense_data:
                        title = expense_data['title']
                        await message.answer(
                            f"Дані про статтю:\nНазва: {title}\nЦе стаття для видалення. Підтвердьте видалення (Так/Ні):"
                        )
                        
                        await state.set_state(DelExpenseForm.waiting_for_confirmation)
                    else:
                        await message.answer("Стаття з таким ID не знайдена.")
                else:
                    await message.answer("Отримано несподіваний формат даних.")
            else:
                await message.answer("Помилка при отриманні даних.")


async def confirm_delete_expense(message: types.Message, state: FSMContext):
    confirmation = message.text.lower()
    data = await state.get_data()
    expense_id = data.get("expense_id")

    if confirmation == "так":
        
        response = requests.delete(f"{REPORT_ALL_URL}/{expense_id}")
        
        if response.status_code == 200:
            await message.answer("Витрата успішно видалена.")
        else:
            await message.answer("Помилка при видаленні витрати.")

    elif confirmation == "ні":
        await message.answer("Видалення витрати відхилено")
    else:
        await message.answer("Буль-ласка, дайте відповідь 'Так' або 'Ні'.")

    
    await message.answer("Оберіть наступну дію:", reply_markup=menu_keyboard)
    await state.clear()




def register_handlers(dp: Dispatcher):
    dp.message(Command("start"))(start_cmd)
    dp.message(lambda message: message.text == "➕ Додати статтю витрат")(add_expense)
    dp.message(lambda message: message.text == "📊 Отримати звіт витрат за вказаний період")(get_report)
    dp.message(lambda message: message.text == "📝 Відредагувати статтю у списку витрат")(edit_expense)
    dp.message(lambda message: message.text == "❌ Видалити статтю у списку витрат")(del_expense)

    
    dp.message(StateFilter(ExpenseForm.waiting_for_title))(get_expense_title)
    dp.message(StateFilter(ExpenseForm.waiting_for_date))(get_expense_date)
    dp.message(StateFilter(ExpenseForm.waiting_for_amount))(get_expense_amount)
    
    dp.message(StateFilter(ReportForm.waiting_for_start_date))(get_start_date)
    dp.message(StateFilter(ReportForm.waiting_for_end_date))(get_end_date)
    
    dp.message(StateFilter(EditExpenseForm.waiting_for_expense_id))(get_expense_id)
    dp.message(StateFilter(EditExpenseForm.waiting_for_new_title))(get_new_title)
    dp.message(StateFilter(EditExpenseForm.waiting_for_new_amount))(get_new_amount)

    dp.message(StateFilter(DelExpenseForm.waiting_for_expense_id))(delete_expense)
    dp.message(StateFilter(DelExpenseForm.waiting_for_confirmation))(confirm_delete_expense)